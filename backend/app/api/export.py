"""
Report Export API - PDF and Excel generation
"""
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from io import BytesIO
from datetime import datetime
from typing import Optional

# PDF generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_RIGHT

# Excel generation
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.core.database import get_db
from app.models.report import Report
from app.models.expense import Expense

from app.models.user import User
from app.models.trip import Trip
from app.core.dependencies import get_current_user

router = APIRouter()

def generate_pdf_report(report: Report, expenses: list, user: User, trip: Trip = None) -> BytesIO:
    """
    Genera un PDF del reporte con todos los gastos
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter,
                           rightMargin=72, leftMargin=72,
                           topMargin=72, bottomMargin=18)
    
    # Container for elements
    elements = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#2563eb'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=12
    )
    
    # Title
    title = Paragraph(f"Reporte de Gastos - {report.title}", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))
    
    # Report info
    info_data = [
        ['Usuario:', user.full_name],
        ['Viaje:', trip.name if trip else 'N/A'],
        ['Período:', f"{report.start_date.strftime('%d/%m/%Y')} - {report.end_date.strftime('%d/%m/%Y')}"],
        ['Generado:', datetime.now().strftime('%d/%m/%Y %H:%M')],
    ]
    
    info_table = Table(info_data, colWidths=[1.5*inch, 4*inch])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#374151')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Summary
    total_amount = sum(e.amount for e in expenses) / 100 if expenses else 0
    summary_heading = Paragraph("Resumen", heading_style)
    elements.append(summary_heading)
    
    summary_data = [
        ['Total de Gastos:', f'{len(expenses)} gastos'],
        ['Monto Total:', f'${total_amount:,.2f}'],
    ]
    
    summary_table = Table(summary_data, colWidths=[1.5*inch, 4*inch])
    summary_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('TEXTCOLOR', (1, -1), (1, -1), colors.HexColor('#059669')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 0.4*inch))
    
    # Expenses table
    expenses_heading = Paragraph("Detalle de Gastos", heading_style)
    elements.append(expenses_heading)
    elements.append(Spacer(1, 0.1*inch))
    
    table_data = [['Fecha', 'Categoría', 'Descripción', 'Comercio', 'Monto']]
    
    for expense in expenses:
        # Manejo de valores nulos y por defecto
        fecha = expense.expense_date.strftime('%d/%m/%Y') if expense.expense_date else 'N/A'
        categoria = expense.category.name if expense.category and expense.category.name else 'N/A'
        descripcion = expense.description or ''
        if len(descripcion) > 30:
            descripcion = descripcion[:30] + '...'
        comercio = expense.merchant or 'N/A'
        if len(comercio) > 20:
            comercio = comercio[:20] + '...'
        monto = expense.amount if expense.amount is not None else 0.0
        monto = float(monto) / 100  # Convertir de centavos a moneda real
        try:
            monto_str = f'${monto:,.2f}'
        except Exception:
            monto_str = '$0.00'
        table_data.append([
            fecha,
            categoria,
            descripcion,
            comercio,
            monto_str
        ])
    
    expenses_table = Table(table_data, colWidths=[1*inch, 1.2*inch, 2*inch, 1.3*inch, 1*inch])
    expenses_table.setStyle(TableStyle([
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        
        # Body
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),
        ('ALIGN', (-1, 1), (-1, -1), 'RIGHT'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
    ]))
    elements.append(expenses_table)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    return buffer


def generate_excel_report(report: Report, expenses: list, user: User) -> BytesIO:
    """
    Genera un Excel del reporte con todos los gastos
    """
    buffer = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Gastos"
    
    # Styles
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    title_font = Font(bold=True, size=16, color="2563EB")
    title_alignment = Alignment(horizontal="center", vertical="center")
    
    border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )
    
    # Title
    ws.merge_cells('A1:E1')
    ws['A1'] = f"Reporte de Gastos - {report.title}"
    ws['A1'].font = title_font
    ws['A1'].alignment = title_alignment
    ws.row_dimensions[1].height = 30
    
    # Info section
    row = 3
    ws[f'A{row}'] = "Usuario:"
    ws[f'B{row}'] = user.full_name
    ws[f'A{row}'].font = Font(bold=True)
    
    row += 1
    ws[f'A{row}'] = "Viaje:"
    ws[f'B{row}'] = report.trip.name if report.trip else 'N/A'
    ws[f'A{row}'].font = Font(bold=True)
    
    row += 1
    ws[f'A{row}'] = "Período:"
    ws[f'B{row}'] = f"{report.start_date.strftime('%d/%m/%Y')} - {report.end_date.strftime('%d/%m/%Y')}"
    ws[f'A{row}'].font = Font(bold=True)
    
    row += 1
    ws[f'A{row}'] = "Generado:"
    ws[f'B{row}'] = datetime.now().strftime('%d/%m/%Y %H:%M')
    ws[f'A{row}'].font = Font(bold=True)
    
    # Summary section
    row += 2
    total_amount = sum(e.amount for e in expenses)
    ws[f'A{row}'] = "Total de Gastos:"
    ws[f'B{row}'] = f"{len(expenses)} gastos"
    ws[f'A{row}'].font = Font(bold=True)
    
    row += 1
    ws[f'A{row}'] = "Monto Total:"
    ws[f'B{row}'] = total_amount
    ws[f'A{row}'].font = Font(bold=True, color="059669")
    ws[f'B{row}'].font = Font(bold=True, size=14, color="059669")
    ws[f'B{row}'].number_format = '"$"#,##0.00'
    
    # Expenses table header
    row += 3
    headers = ['Fecha', 'Categoría', 'Descripción', 'Comercio', 'Monto']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Expenses data
    for expense in expenses:
        row += 1
        ws.cell(row=row, column=1, value=expense.expense_date.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=2, value=expense.category.name if expense.category else 'N/A')
        ws.cell(row=row, column=3, value=expense.description)
        ws.cell(row=row, column=4, value=expense.merchant or 'N/A')
        ws.cell(row=row, column=5, value=expense.amount)
        
        # Formatting
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            
        # Amount formatting
        ws.cell(row=row, column=5).number_format = '"$"#,##0.00'
        ws.cell(row=row, column=5).alignment = Alignment(horizontal="right", vertical="center")
    
    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 12
    
    wb.save(buffer)
    buffer.seek(0)
    return buffer



@router.get("/{report_id}/export/pdf")
async def export_report_pdf(
    report_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Exportar reporte a PDF
    """
    report = db.query(Report).filter(Report.id == report_id).first()
    if not report:
        raise HTTPException(status_code=404, detail="Reporte no encontrado")
    # Verificar permisos
    if report.user_id != current_user.id and current_user.role not in ["admin", "manager"]:
        raise HTTPException(status_code=403, detail="No tienes permiso para ver este reporte")
    # Obtener gastos del reporte
    expenses = db.query(Expense).filter(
        Expense.report_id == report_id
    ).all()
    # Obtener el trip asociado (si existe) a partir del primer gasto
    trip = None
    for expense in expenses:
        if expense.trip_id:
            trip = db.query(Trip).filter(Trip.id == expense.trip_id).first()
            break
    # Generar PDF
    pdf_buffer = generate_pdf_report(report, expenses, current_user, trip)
    filename = f"reporte_{report.id}_{datetime.now().strftime('%Y%m%d')}.pdf"
    return StreamingResponse(
        pdf_buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/{report_id}/export/excel")
async def export_report_excel(
    report_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Exportar reporte a Excel
    """
    report = db.query(Report).filter(Report.id == report_id).first()
    
    if not report:
        raise HTTPException(status_code=404, detail="Reporte no encontrado")
    
    # Verificar permisos
    if report.user_id != current_user.id and current_user.role not in ["admin", "manager"]:
        raise HTTPException(status_code=403, detail="No tienes permiso para ver este reporte")
    
    # Obtener gastos del reporte
    expenses = db.query(Expense).filter(
        Expense.report_id == report_id
    ).all()
    
    # Generar Excel
    excel_buffer = generate_excel_report(report, expenses, current_user)
    
    filename = f"reporte_{report.id}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        excel_buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
